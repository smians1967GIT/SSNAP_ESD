import pandas as pd
import gradio as gr
import os
import tempfile
from openpyxl import load_workbook

def extract_esd_metrics_with_full_labels(file_obj):
    if file_obj is None:
        return "❌ No file uploaded", None, None

    try:
        xls = pd.ExcelFile(file_obj.name)
        wb = load_workbook(file_obj.name, data_only=True)
        target_sheets = [s for s in xls.sheet_names if s.startswith("L")]
    except Exception as e:
        return f"❌ Failed to read Excel: {e}", None, None

    combined_records = []

    for sheet in target_sheets:
        try:
            df = xls.parse(sheet, header=None)
            ws = wb[sheet]

            # Extract team metadata from rows 0–3 and columns F onwards
            team_data = []
            for col in range(5, df.shape[1]):
                team_type = df.iloc[0, col]
                region = df.iloc[1, col]
                trust = df.iloc[2, col]
                esd_team = df.iloc[3, col]

                if pd.notna(esd_team):
                    team_data.append({
                        "col": col,
                        "Team Type": str(team_type).strip() if pd.notna(team_type) else None,
                        "Region": str(region).strip() if pd.notna(region) else None,
                        "Trust": str(trust).strip() if pd.notna(trust) else None,
                        "ESD Team": str(esd_team).strip()
                    })

            # ✅ Build metric label mapping with forward-fill logic
            metric_labels = {}
            last_label = None
            for row in range(5, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if val and str(val).strip().lower() != "nan":
                    last_label = str(val).replace("\n", " ").strip()
                metric_labels[row] = last_label

            # Loop through each row for % or median metrics
            for i in range(4, df.shape[0]):
                raw_type = df.iloc[i, 3]
                data_type = str(raw_type).strip().lower() if pd.notna(raw_type) else ""

                if "%" in data_type or "median" in data_type:
                    metric_id = df.iloc[i, 1]
                    row_excel = i + 2  # Excel row alignment
                    metric_label = metric_labels.get(row_excel, "")

                    for team in team_data:
                        value = df.iloc[i, team["col"]]
                        if pd.notna(value) and str(value).strip() not in ["Too few to report", ".", "N/A", ""]:
                            combined_records.append({
                                **team,
                                "Sheet": sheet,
                                "Metric ID": metric_id,
                                "Metric Label": metric_label,
                                "Data Type": df.iloc[i, 3],
                                "Value": value
                            })

        except Exception as e:
            print(f"❌ Error processing sheet {sheet}: {e}")
            continue

    if not combined_records:
        return "❌ No metrics with '%' or 'median' found.", None, None

    df_final = pd.DataFrame(combined_records)
    out_path = os.path.join(tempfile.gettempdir(), "SSNAP_ESD_Metrics_Complete.xlsx")
    df_final.to_excel(out_path, index=False)

    return f"✅ Extracted {len(df_final)} metric rows with full labels.", df_final, out_path

# ---------- Gradio UI ----------
with gr.Blocks() as demo:
    gr.Markdown("### 📊 SSNAP ESD Metric Extractor\nUpload a SSNAP Excel file to extract metrics containing **%** or **median** with **full metric labels** from `L*` sheets.")

    file_input = gr.File(label="📁 Upload SSNAP Excel File (.xlsx)", file_types=[".xlsx"])
    status = gr.Textbox(label="Status", interactive=False)
    table = gr.Dataframe(label="📋 Extracted Metrics Table")
    download = gr.File(label="⬇️ Download Extracted Excel File")

    file_input.change(
        fn=extract_esd_metrics_with_full_labels,
        inputs=[file_input],
        outputs=[status, table, download]
    )

if __name__ == "__main__":
    demo.launch()
